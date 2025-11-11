import sys
import subprocess
import importlib

# --- Dependency Check and Auto-Install ---
required_packages = {
    "openpyxl": "openpyxl",
    "tkinter": "tk",  # tkinter included in Python/Anaconda
    "customtkinter": "customtkinter"
}

missing_packages = []

for package, install_name in required_packages.items():
    try:
        importlib.import_module(package)
    except ImportError:
        if package == "tkinter":
            missing_packages.append(package)
        else:
            try:
                print(f"Installing missing package: {install_name}...")
                subprocess.check_call([sys.executable, "-m", "pip", "install", install_name])
            except Exception as e:
                missing_packages.append(package)

if missing_packages:
    import tkinter
    from tkinter import messagebox
    msg = "The following required packages could not be installed:\n" + "\n".join(missing_packages)
    msg += "\n\nPlease install them manually and rerun the application."
    root = tkinter.Tk()
    root.withdraw()
    messagebox.showerror("Missing Dependencies", msg)
    sys.exit(1)

# --- All dependencies available ---
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class MasterIndexBuilder(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel Master Index Builder")
        self.geometry("650x450")
        self.resizable(False, False)

        self.folder_path = None

        # --- UI Elements ---
        self.label_title = ctk.CTkLabel(self, text="Excel Master Index Builder", font=("Segoe UI", 20, "bold"))
        self.label_title.pack(pady=10)

        self.btn_select = ctk.CTkButton(self, text="Select Folder", command=self.select_folder)
        self.btn_select.pack(pady=10)

        self.label_path = ctk.CTkLabel(self, text="No folder selected", wraplength=600, text_color="gray")
        self.label_path.pack(pady=5)

        self.progress_bar = ctk.CTkProgressBar(self, width=500)
        self.progress_bar.set(0)
        self.progress_bar.pack(pady=10)

        self.text_output = ctk.CTkTextbox(self, width=600, height=220)
        self.text_output.pack(pady=10)

        self.btn_start = ctk.CTkButton(self, text="Build Master Index", state="disabled", command=self.start_process)
        self.btn_start.pack(pady=10)

    # --- Folder Selection ---
    def select_folder(self):
        folder = filedialog.askdirectory(title="Select Folder Containing Excel Files")
        if folder:
            self.folder_path = Path(folder)
            self.label_path.configure(text=str(folder))
            self.btn_start.configure(state="normal")

    # --- Start threaded process ---
    def start_process(self):
        self.btn_start.configure(state="disabled")
        self.text_output.delete("1.0", "end")
        thread = threading.Thread(target=self.build_master_index)
        thread.start()

    # --- Build Master Index ---
    def build_master_index(self):
        folder_path = self.folder_path
        output_file = folder_path / "Master_Index.xlsx"
        wb_index = Workbook()
        ws_index = wb_index.active
        ws_index.title = "Master Index"

        headers = ["File Name", "Sheet Name", "Link Formula"]
        ws_index.append(headers)
        for col in range(1, len(headers)+1):
            ws_index.cell(row=1, column=col).font = Font(bold=True)

        row = 2
        excel_files = [f for f in folder_path.glob("*.xls*") if f.name != "Master_Index.xlsx"]
        total_files = len(excel_files)
        report = {"processed":0, "errors":0, "details":[]}

        for idx, file_path in enumerate(excel_files, start=1):
            self.text_output.insert("end", f"[{idx}/{total_files}] {file_path.name} ... ")
            self.text_output.update()
            try:
                wb = load_workbook(file_path, read_only=True)
                for sheet_name in wb.sheetnames:
                    ws_index.cell(row=row, column=1, value=file_path.name)
                    ws_index.cell(row=row, column=2, value=sheet_name)

                    full_path = str(file_path.resolve()).replace('"', '""')
                    sheet_name_escaped = sheet_name.replace("'", "''")
                    formula = f'=HYPERLINK("{full_path}#\'{sheet_name_escaped}\'!A1", "Open Sheet")'
                    ws_index.cell(row=row, column=3, value=formula)
                    row +=1

                wb.close()
                report["processed"] += 1
                report["details"].append(f"✅ {file_path.name} — {len(wb.sheetnames)} sheets indexed")
                self.text_output.insert("end", "✅ Done\n")
            except Exception as e:
                report["errors"] += 1
                report["details"].append(f"❌ Error reading {file_path.name}: {e}")
                self.text_output.insert("end", "❌ Failed\n")
            self.progress_bar.set(idx / total_files)
            self.text_output.update()

        # Adjust column widths
        for col in range(1, 4):
            max_len = max(len(str(c.value)) if c.value else 0 for c in ws_index[get_column_letter(col)])
            ws_index.column_dimensions[get_column_letter(col)].width = max_len + 3

        wb_index.save(output_file)
        self.show_summary(output_file, report)
        self.btn_start.configure(state="normal")

    # --- Display Summary ---
    def show_summary(self, output_file, report):
        self.text_output.insert("end", "\n" + "="*60 + "\n")
        self.text_output.insert("end", f"📊 PROCESS REPORT\n")
        self.text_output.insert("end", "="*60 + "\n")
        self.text_output.insert("end", f"Total files processed : {report['processed']}\n")
        self.text_output.insert("end", f"Files with errors      : {report['errors']}\n")
        self.text_output.insert("end", f"Output file created at : {output_file}\n\n")
        self.text_output.insert("end", "Detailed log:\n")
        for line in report["details"]:
            self.text_output.insert("end", f" - {line}\n")
        self.text_output.insert("end", "="*60 + "\n")
        self.text_output.update()
        messagebox.showinfo("Process Complete", f"Master Index created at:\n{output_file}")

if __name__ == "__main__":
    app = MasterIndexBuilder()
    app.mainloop()
